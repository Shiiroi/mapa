#!/usr/bin/env python3
"""Build data-sets/data/clean/popcen_2010_2024.csv from the two PSA workbooks.

Spine (PSGC code + name + level + 2024 population, down to barangay) comes from
the PSGC publication datafile. The 2010/2015/2020 census counts (down to
city/municipality) come from Table B, whose rows are name-keyed only, so we match
them back to PSGC codes by region + normalized name, falling back to an exact
2024-population match within the region. Country totals are summed from regions.

Run:  scripts/py/.venv/bin/python scripts/py/convert_population.py
"""

import csv
import os
import re
import sys
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
FRONTEND = os.path.abspath(os.path.join(HERE, "..", ".."))
SOURCE = os.path.join(FRONTEND, "data-sets", "source")
OUT = os.path.join(FRONTEND, "data-sets", "data", "clean", "popcen_2010_2024.csv")

PSGC_XLSX = os.path.join(SOURCE, "PSGC-1Q-2026-Publication-Datafile-2.xlsx")
TABLEB_XLSX = os.path.join(
    SOURCE,
    "3_Table B - Population and Annual PGR by Province, City, and Municipality - By Region - rev_0.xlsx",
)

# PSGC geographic-level code -> division_stats level string.
LEVEL_MAP = {"Reg": "region", "Prov": "province", "City": "municipality", "Mun": "municipality", "Bgy": "barangay"}

ABBREV = {
    "STO": "SANTO",
    "STA": "SANTA",
    "STO.": "SANTO",
    "STA.": "SANTA",
    "GEN": "GENERAL",
    "MT": "MOUNT",
}


def norm(name: str) -> str:
    """Normalize a place name for matching: upper, strip 'CITY OF'/footnotes/punct/Ñ."""
    if name is None:
        return ""
    s = str(name).upper().strip()
    s = s.replace("Ñ", "N").replace("Ã'", "N")
    s = re.sub(r"\(CAPITAL\)", " ", s)
    s = re.sub(r"\(.*?\)", " ", s)            # drop parentheticals
    s = re.sub(r"\bCITY OF\b", " ", s)
    s = re.sub(r"\bCITY\b", " ", s)
    s = re.sub(r"[^A-Z0-9 ]", " ", s)         # drop *, **, /, ., - and other footnote marks
    toks = [ABBREV.get(t, t) for t in s.split()]
    # drop lone-digit footnote markers (e.g. "MAKATI 1", "NEGROS OCCIDENTAL 2")
    toks = [t for t in toks if not t.isdigit()]
    return " ".join(toks).strip()


def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None


def load_psgc():
    """Return spine rows + lookup indices from the PSGC sheet."""
    wb = openpyxl.load_workbook(PSGC_XLSX, read_only=True, data_only=True)
    ws = wb["PSGC"]
    spine = {}                                  # code -> {code, level, name, pop_2024}
    idx = {
        "region_by_norm": {},                   # norm -> region code
        "prov_by_region": defaultdict(dict),    # r2 -> {norm: code}
        "citymun_by_region": defaultdict(dict), # r2 -> {norm: [codes]}
        "pop24_by_region": defaultdict(dict),   # r2 -> {pop_2024: code}  (city/mun)
        "prov_by_norm": defaultdict(list),      # norm -> [codes]  (all regions)
        "citymun_by_norm": defaultdict(list),   # norm -> [codes]  (all regions)
        "prov_by_pop24": {},                    # pop_2024 -> code  (provinces, all regions)
        "citymun_by_pop24": defaultdict(list),  # pop_2024 -> [codes]  (city/mun, all regions)
    }
    for r in ws.iter_rows(min_row=2, values_only=True):
        code, name, lvl, pop = r[0], r[1], r[3], r[8]
        if code is None or lvl not in LEVEL_MAP:
            continue
        code = str(code).strip()
        pop24 = to_int(pop)
        spine[code] = {"code": code, "level": LEVEL_MAP[lvl], "name": str(name).strip(), "pop_2024": pop24}
        r2 = code[:2]
        n = norm(name)
        if lvl == "Reg":
            idx["region_by_norm"][n] = code
        elif lvl == "Prov":
            idx["prov_by_region"][r2][n] = code
            idx["prov_by_norm"][n].append(code)
            if pop24 is not None:
                idx["prov_by_pop24"].setdefault(pop24, code)
        elif lvl in ("City", "Mun"):  # only city/municipality — never barangays/sub-muns
            idx["citymun_by_region"][r2].setdefault(n, []).append(code)
            idx["citymun_by_norm"][n].append(code)
            if pop24 is not None:
                idx["pop24_by_region"][r2].setdefault(pop24, code)
                idx["citymun_by_pop24"][pop24].append(code)
    wb.close()
    return spine, idx


def main():
    spine, idx = load_psgc()
    region_by_norm = idx["region_by_norm"]
    prov_by_region = idx["prov_by_region"]
    citymun_by_region = idx["citymun_by_region"]
    pop24_by_region = idx["pop24_by_region"]

    # extra[code] = (pop_2010, pop_2015, pop_2020) overlaid onto the spine
    extra = {}
    region_year_totals = defaultdict(lambda: [0, 0, 0, 0])  # year-index totals for country
    unmatched = []

    wb = openpyxl.load_workbook(TABLEB_XLSX, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, max_col=6, values_only=True))
        region_code = None
        r2 = None
        cur_prov = None  # 4-digit prefix of the current province block
        for row in rows:
            name = row[0]
            if not name or not str(name).strip():
                continue
            vals = [to_int(row[2]), to_int(row[3]), to_int(row[4]), to_int(row[5])]
            n = norm(name)

            # region total row: first row whose name resolves to a PSGC region
            if region_code is None:
                if n in region_by_norm:
                    region_code = region_by_norm[n]
                    r2 = region_code[:2]
                    if all(v is not None for v in vals):
                        extra[region_code] = (vals[0], vals[1], vals[2])
                        for i in range(4):
                            if vals[i] is not None:
                                region_year_totals[i]  # touch
                continue

            if not any(v is not None for v in vals):
                continue

            # Province header row. A province row carries the province's own 2024 total
            # (PSGC province pop_2024 == Table B province total, verified), which is what
            # separates it from same-named cities/municipalities ("City of Cebu" -> "CEBU",
            # municipality "Quezon" -> province "Quezon", …). Cross-region provinces whose
            # PSGC region differs from the Table B sheet (e.g. Sulu) match by 2024 pop.
            prov_code = None
            if vals[3] is not None:
                pc = prov_by_region[r2].get(n)
                if pc is not None and spine[pc]["pop_2024"] == vals[3]:
                    prov_code = pc
                if prov_code is None:
                    pc2 = idx["prov_by_pop24"].get(vals[3])
                    if pc2 is not None and norm(spine[pc2]["name"]) == n:
                        prov_code = pc2
            if prov_code is not None:
                cur_prov = prov_code[:4]
                extra[prov_code] = (vals[0], vals[1], vals[2])
                continue

            # City / municipality row. Exact 2024 population is the authoritative shared
            # key between the two sources, so match on it first (disambiguating rare
            # pop ties by name then region), then fall back to name within the region.
            code = None
            pop_cands = idx["citymun_by_pop24"].get(vals[3], []) if vals[3] is not None else []
            if len(pop_cands) == 1:
                code = pop_cands[0]
            elif len(pop_cands) > 1:
                named = [c for c in pop_cands if norm(spine[c]["name"]) == n]
                in_reg = [c for c in pop_cands if c[:2] == r2]
                if len(named) == 1:
                    code = named[0]
                elif len(in_reg) == 1:
                    code = in_reg[0]
            if code is None:                                # name within region
                cands = citymun_by_region[r2].get(n, [])
                if len(cands) == 1:
                    code = cands[0]
                elif len(cands) > 1 and cur_prov:
                    same = [c for c in cands if c[:4] == cur_prov]
                    code = same[0] if len(same) == 1 else None
            if code is None:                                # unique global name
                gn = idx["citymun_by_norm"].get(n, [])
                if len(gn) == 1:
                    code = gn[0]
            if code is None:
                unmatched.append((sheet, str(name).strip(), vals[3]))
                continue
            extra[code] = (vals[0], vals[1], vals[2])
    wb.close()

    # country totals = sum of region rows (2010/2015/2020 from Table B, 2024 from spine sum)
    reg_codes = [c for c, v in spine.items() if v["level"] == "region"]
    country = [0, 0, 0]
    for c in reg_codes:
        e = extra.get(c)
        if e:
            for i in range(3):
                country[i] += e[i] or 0
    pop24_country = sum(spine[c]["pop_2024"] or 0 for c in reg_codes)
    spine["0000000000"] = {"code": "0000000000", "level": "country", "name": "Philippines", "pop_2024": pop24_country}
    extra["0000000000"] = (country[0], country[1], country[2])

    # write combined CSV ordered by psgc
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    n_written = 0
    with_all_years = 0
    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["psgc", "level", "name", "pop_2010", "pop_2015", "pop_2020", "pop_2024"])
        for code in sorted(spine):
            s = spine[code]
            e = extra.get(code, (None, None, None))
            row = [code, s["level"], s["name"], e[0], e[1], e[2], s["pop_2024"]]
            w.writerow(["" if v is None else v for v in row])
            n_written += 1
            if all(v is not None for v in (e[0], e[1], e[2], s["pop_2024"])):
                with_all_years += 1

    # report
    by_level = defaultdict(int)
    filled = defaultdict(int)
    for code, s in spine.items():
        by_level[s["level"]] += 1
        if extra.get(code, (None,))[0] is not None:
            filled[s["level"]] += 1
    print(f"Wrote {n_written} rows -> {os.path.relpath(OUT, FRONTEND)}")
    print(f"  rows with all four years: {with_all_years}")
    for lvl in ("country", "region", "province", "municipality", "barangay"):
        print(f"  {lvl:13s} total={by_level[lvl]:6d}  has 2010-2020={filled[lvl]:6d}")
    print(f"\nUnmatched Table B rows: {len(unmatched)}")
    for sheet, name, p24 in unmatched[:60]:
        print(f"  [{sheet}] {name!r}  2024={p24}")
    if len(unmatched) > 60:
        print(f"  … and {len(unmatched) - 60} more")
    return 0 if len(unmatched) == 0 else 0


if __name__ == "__main__":
    sys.exit(main())
