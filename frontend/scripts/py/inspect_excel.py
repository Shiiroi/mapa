import openpyxl
import os

excel_path = "/Users/vince/personal-projects/mapa/mapa/frontend/data-sets/source/4_Household Population by Age Group and Sex_Philippines_2020 CPH_rev.xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True)
print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    # Search for Quezon City in first few columns
    row_idx = 0
    for row in ws.iter_rows(max_row=3000, max_col=10, values_only=True):
        row_idx += 1
        for cell in row:
            if cell and "Quezon City" in str(cell):
                print(f"Found in {sheetname} at row {row_idx}: {row}")
