#!/usr/bin/env python3
"""Extract 2020 CPH household population by age group and sex from PSA xlsx."""

from __future__ import annotations

import csv
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = SCRIPT_DIR / "../../public"
XLSX_GLOB = "4_Household Population by Age Group and Sex_Philippines_2020 CPH_rev.xlsx"
OUT_CSV = PUBLIC_DIR / "household_agesex_2020_raw.csv"

TAB_TO_REGION: dict[str, str] = {
    "Philippines": "PHILIPPINES",
    "NCR": "NATIONAL CAPITAL REGION (NCR)",
    "CAR": "CORDILLERA ADMINISTRATIVE REGION (CAR)",
    "Reg I": "REGION I (ILOCOS REGION)",
    "Reg II": "REGION II (CAGAYAN VALLEY)",
    "Reg III": "REGION III (CENTRAL LUZON)",
    "CALABARZON": "REGION IV-A (CALABARZON)",
    "MIMAROPA": "MIMAROPA REGION",
    "Reg V": "REGION V (BICOL REGION)",
    "Reg VI": "REGION VI (WESTERN VISAYAS)",
    "Reg VII": "REGION VII (CENTRAL VISAYAS)",
    "Reg VIII": "REGION VIII (EASTERN VISAYAS)",
    "Reg IX": "REGION IX (ZAMBOANGA PENINSULA)",
    "Reg X": "REGION X (NORTHERN MINDANAO)",
    "Reg XI": "REGION XI (DAVAO REGION)",
    "Reg XII": "REGION XII (SOCCSKSARGEN)",
    "CARAGA": "REGION XIII (CARAGA)",
    "BARMM": "BANGSAMORO AUTONOMOUS REGION IN MUSLIM MINDANAO (BARMM)",
}

AGE_GROUP_RE = re.compile(
    r"^(Total|0\s*-\s*4|5\s*-\s*9|10\s*-\s*14|15\s*-\s*19|20\s*-\s*24|25\s*-\s*29|"
    r"30\s*-\s*34|35\s*-\s*39|40\s*-\s*44|45\s*-\s*49|50\s*-\s*54|55\s*-\s*59|"
    r"60\s*-\s*64|65\s*-\s*69|70\s*-\s*74|75\s*-\s*79|80 years and over)$",
    re.I,
)

COLUMNS = [
    "level",
    "region",
    "province",
    "name",
    "age_group",
    "both_sexes",
    "male",
    "female",
]

EXPECTED_AGE_ROWS = 18  # Total + 17 age bands


def norm_place(raw: object) -> str:
  """Normalize a place name for cross-referencing against PSGC (accent/case/suffix-insensitive)."""
  s = unicodedata.normalize("NFD", str(raw))
  s = "".join(c for c in s if unicodedata.category(c) != "Mn")
  s = s.upper().strip()
  s = re.sub(r"\s+", " ", s)
  s = re.sub(r"\bSTO\.?\s+", "SANTO ", s)
  s = re.sub(r"\bSTA\.?\s+", "SANTA ", s)
  s = re.sub(r"\bGEN\.?\s+", "GENERAL ", s)
  s = re.sub(r"\bMT\.?\s+", "MOUNT ", s)
  s = re.sub(r"\bDR\.?\s+", "DOCTOR ", s)
  s = re.sub(r"^CITY OF ", "", s)
  s = re.sub(r"^MUNICIPALITY OF ", "", s)
  s = re.sub(r"\s+CITY$", "", s)
  s = re.sub(r"\s*\([^)]*\)\s*$", "", s)  # drop trailing "(Capital)", "(excluding ...)", aliases
  return re.sub(r"\s+", " ", s).strip()


class PsgcHierarchy:
  """PSGC province lookup plus each province's set of municipality/city names.

  Used to tell a real province header apart from a municipality that merely shares
  a province's name (e.g. the town of Rizal in Cagayan vs. Rizal province)."""

  def __init__(self) -> None:
    self.province_name_to_code: dict[str, str] = {}  # norm name -> 5-digit province code
    self.province_munis: dict[str, set[str]] = defaultdict(set)  # 5-digit code -> norm muni/city names

  def is_municipality_of(self, prov_code: str | None, norm: str) -> bool:
    return bool(prov_code) and norm in self.province_munis.get(prov_code, set())


def load_psgc_hierarchy() -> PsgcHierarchy:
  h = PsgcHierarchy()
  psgc_path = PUBLIC_DIR / "psgc.csv"
  if not psgc_path.exists():
    return h
  mun_rows: list[tuple[str, str]] = []
  with psgc_path.open(encoding="utf-8", errors="replace") as f:
    reader = csv.DictReader(f)
    for row in reader:
      lvl = row.get("Geographic Level", "").strip()
      code = row.get("10-digit PSGC", "").strip().rjust(10, "0")
      name = row.get("Name", "").strip()
      if not name or len(code) < 5:
        continue
      if lvl == "Prov":
        h.province_name_to_code[norm_place(name)] = code[:5]
      elif lvl in ("Mun", "City"):
        mun_rows.append((code, name))
  for code, name in mun_rows:
    h.province_munis[code[:5]].add(norm_place(name))
  return h


def clean_name(raw: str) -> str:
  s = raw.strip()
  s = re.sub(r"\s+\d+$", "", s)  # trailing footnote numbers
  s = re.sub(r"\s+\*$", "", s)
  s = re.sub(r"\s+\*\*$", "", s)
  return s.strip()


def parse_int(val: object) -> int | None:
  if val is None:
    return None
  if isinstance(val, (int, float)):
    return int(val)
  s = str(val).strip().replace(",", "")
  if not s or s == "-":
    return None
  try:
    return int(float(s))
  except ValueError:
    return None


def is_region_name(name: str, tab_region: str) -> bool:
  upper = name.upper()
  if upper.startswith("NATIONAL CAPITAL") or upper.startswith("CORDILLERA ADMINISTRATIVE"):
    return True
  if upper.startswith("REGION ") or ("REGION" in upper and upper.endswith(")")):
    return True
  return upper == tab_region.upper()


def classify_level(
  name: str,
  tab_region: str,
  hierarchy: PsgcHierarchy,
  current_prov_code: str | None,
) -> tuple[str, str | None]:
  """Return (level, province_code). A name is a province only if it matches a PSGC province
  and is not a municipality of the province block we are currently inside."""
  upper = name.upper()
  if upper == "PHILIPPINES":
    return "country", None
  if is_region_name(name, tab_region):
    return "region", None
  norm = norm_place(name)
  prov_code = hierarchy.province_name_to_code.get(norm)
  if prov_code and not hierarchy.is_municipality_of(current_prov_code, norm):
    return "province", prov_code
  if upper.startswith("CITY OF ") or upper.endswith(" CITY") or "CITY" in upper:
    return "city", current_prov_code
  return "municipality", current_prov_code


def is_place_header(a: str | None, b: object) -> bool:
  if not a or not str(a).strip():
    return False
  if b is not None and str(b).strip() not in ("", "None"):
    return False
  if AGE_GROUP_RE.match(str(a).strip()):
    return False
  if str(a).strip().lower().startswith("age group"):
    return False
  if "Household Population" in str(a):
    return False
  return True


def find_xlsx() -> Path:
  matches = list(PUBLIC_DIR.glob(XLSX_GLOB))
  if not matches:
    matches = list(PUBLIC_DIR.glob("4_Household*.xlsx"))
  if not matches:
    print(f"ERROR: missing household xlsx in {PUBLIC_DIR}", file=sys.stderr)
    sys.exit(1)
  return matches[0]


def extract_sheet(ws, tab_name: str, hierarchy: PsgcHierarchy) -> tuple[list[dict], list[str]]:
  tab_region = TAB_TO_REGION.get(tab_name, tab_name)
  rows_out: list[dict] = []
  errors: list[str] = []

  current_name = ""
  current_level = ""
  current_province = ""
  current_prov_code: str | None = None
  current_region = tab_region if tab_name != "Philippines" else ""
  block_rows: list[tuple[str, int | None, int | None, int | None]] = []

  def flush_block() -> None:
    nonlocal block_rows
    if not current_name or not block_rows:
      block_rows = []
      return
    if len(block_rows) != EXPECTED_AGE_ROWS:
      errors.append(
        f"[{tab_name}] {current_name}: expected {EXPECTED_AGE_ROWS} rows, got {len(block_rows)}"
      )
    for age_group, both_v, male_v, female_v in block_rows:
      rows_out.append(
        {
          "level": current_level,
          "region": current_region,
          "province": current_province if current_level in ("municipality", "city") else "",
          "name": current_name,
          "age_group": age_group,
          "both_sexes": both_v,
          "male": male_v,
          "female": female_v,
        }
      )
    block_rows = []

  for row in ws.iter_rows(values_only=True):
    a = row[0] if len(row) > 0 else None
    b = row[1] if len(row) > 1 else None
    c = row[2] if len(row) > 2 else None
    d = row[3] if len(row) > 3 else None

    if a is None and b is None:
      continue

    a_str = str(a).strip() if a is not None else ""

    if is_place_header(a_str, b):
      flush_block()
      current_name = clean_name(a_str)
      if tab_name == "Philippines" and current_name.upper() == "PHILIPPINES":
        current_level = "country"
        current_region = ""
        current_province = ""
        current_prov_code = None
      elif tab_name == "Philippines":
        current_level = "region"
        current_region = current_name
        current_province = ""
        current_prov_code = None
      else:
        current_level, prov_code = classify_level(
          current_name, tab_region, hierarchy, current_prov_code
        )
        if current_level == "region":
          current_region = current_name
          current_province = ""
          current_prov_code = None
        elif current_level == "province":
          current_province = current_name
          current_prov_code = prov_code
        # city/municipality keep the current province context
      continue

    if a_str and AGE_GROUP_RE.match(a_str):
      both_v = parse_int(b)
      male_v = parse_int(c)
      female_v = parse_int(d)
      block_rows.append((a_str if a_str != "Total" else "Total", both_v, male_v, female_v))

  flush_block()
  return rows_out, errors


def main() -> None:
  xlsx_path = find_xlsx()
  hierarchy = load_psgc_hierarchy()
  wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

  all_rows: list[dict] = []
  all_errors: list[str] = []

  for tab_name in wb.sheetnames:
    ws = wb[tab_name]
    sheet_rows, errors = extract_sheet(ws, tab_name, hierarchy)
    places = len({(r["name"], r["level"]) for r in sheet_rows})
    print(f"  {tab_name}: {places} places, {len(sheet_rows)} age rows")
    all_rows.extend(sheet_rows)
    all_errors.extend(errors)

  OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
  with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=COLUMNS)
    w.writeheader()
    w.writerows(all_rows)

  print(f"Wrote {len(all_rows)} rows -> {OUT_CSV.relative_to(PUBLIC_DIR.parent)}")
  if all_errors:
    print(f"WARN: {len(all_errors)} block issues:", file=sys.stderr)
    for e in all_errors[:20]:
      print(f"  {e}", file=sys.stderr)
    if len(all_errors) > 20:
      print(f"  ... and {len(all_errors) - 20} more", file=sys.stderr)


if __name__ == "__main__":
  main()
